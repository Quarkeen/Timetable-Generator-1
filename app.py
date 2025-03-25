from flask import Flask, render_template, request, redirect, url_for
import os
import random
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

timetable_context = {}

# ------------------------------
# SCHEDULING LOGIC
# ------------------------------
def parse_credits(credit_str):
    """
    Parse a credit string of the form "L-T-P-S-C" and return (L, T, P, S, C).
    If parsing fails, return zeros.
    """
    try:
        L_str, T_str, P_str, S_str, C_str = credit_str.split('-')
        return int(L_str), int(T_str), int(P_str), int(S_str), int(C_str)
    except:
        return 0, 0, 0, 0, 0

def schedule_courses(courses, color_map):
    """
    Randomly schedule each course's L (lecture), T (tutorial), and P (lab) hours
    across 5 days (MON-FRI) with distinct time slots for lectures, tutorials, labs,
    and non-schedulable breaks.
    
    Time slots:
      9:00 - 10:30 AM  (Lecture slot)
      10:30 - 11:00 AM (Morning Break, non-schedulable)
      11:00 - 12:30 PM (Lecture slot)
      12:30 - 1:30 PM  (Lecture/Tutorial slot)
      1:30 - 2:30 PM   (Lunch Break, non-schedulable)
      2:30 - 4:00 PM   (Lecture slot) or used for lab replacement
      5:00 - 6:30 PM   (Reserved for HS205)
    """

    # Days (Monday to Friday)
    days = ["MON", "TUE", "WED", "THU", "FRI"]

    # Define slots: list of all slot labels in order.
    all_slots = [
        "8:00 - 9:00 AM(Minor Slot)",  # Minor Slot
        "9:00 - 10:30 AM",      # Lecture slot
        "10:30 - 11:00 AM",     # Morning Break (non-schedulable)
        "11:00 - 12:30 PM",     # Lecture slot
        "12:30 - 1:30 PM",      # Lecture/Tutorial slot
        "1:30 - 2:30 PM",       # Lunch Break (non-schedulable)
        "2:30 - 4:30 PM",       # Lecture slot OR lab slot (if lab assigned, we mark it specially)
        "5:00 - 6:30 PM",        # Reserved for HS205
        "6:30 - 8:30 PM(Minor Slot)"  # Minor Slot
    ]

    # For random allocation, define separately:
    lecture_slots = ["9:00 - 10:30 AM", "11:00 - 12:30 PM"]
    tutorial_slots = ["12:30 - 1:30 PM"]  # We will mark tutorial as course+"_TUT" in this slot
    lab_slots = ["2:30 - 4:30 PM"]  # We will mark lab as course+"_LAB(2hrs)" in this slot

    # Initialize timetable
    timetable = {}
    for d in days:
        timetable[d] = {}
        for slot in all_slots:
            if slot == "10:30 - 11:00 AM":
                timetable[d][slot] = "Morning Break"
            elif slot == "1:30 - 2:30 PM":
                timetable[d][slot] = "Lunch Break"
            else:
                timetable[d][slot] = ""

    # Helper function: check if a course (or course variant) is already scheduled in a day.
    def course_in_day(course_label, day):
        for val in timetable[day].values():
            if val == course_label:
                return True
        return False

    MAX_ATTEMPTS = 50  # maximum attempts per course scheduling loop

    # 1) Reserve HS205 in slot "5:00 - 6:30 PM" on a fixed day (FRI)
    for course in courses:
        if course.get("Course Code", "").strip().upper() == "HS205":
            timetable["FRI"]["5:00 - 6:30 PM"] = "HS205"
            timetable["WED"]["5:00 - 6:30 PM"] = "HS205"
            break

    # 2) Schedule lab sessions for courses with P > 0.
    for course in courses:
        code_raw = course.get("Course Code", "")
        code = str(code_raw).strip()
        _, _, P, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        if P > 0:
            labs_needed = P // 2 if P >= 2 else P  # each lab slot covers 2 hours
            attempts = 0
            while labs_needed > 0 and attempts < MAX_ATTEMPTS:
                attempts += 1
                random.shuffle(days)
                for d in days:
                    # Avoid duplicating lab on same day
                    if course_in_day(code + "_LAB", d) or course_in_day(code, d):
                        continue
                    for slot in lab_slots:
                        if timetable[d][slot] == "":
                            timetable[d][slot] = code + "_LAB(2hrs)"
                            labs_needed -= 1
                            break
                    if labs_needed <= 0:
                        break
            # End lab allocation loop for this course

    # 3) Schedule lecture sessions (L hours).
    for course in courses:
        code_raw = course.get("Course Code", "")
        code = str(code_raw).strip()
        if code.upper() == "HS205":
            continue  # already scheduled
        L, _, _, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        lectures_needed = L-1
        attempts = 0
        while lectures_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(lecture_slots)
            placed = False
            for d in days:
                if course_in_day(code, d):
                    continue  # already scheduled that day for lecture
                for slot in lecture_slots:
                    if timetable[d][slot] == "":
                        timetable[d][slot] = code
                        lectures_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break  # if no placement in one full iteration, break to avoid infinite loop

    # 4) Schedule tutorial sessions (T hours).
    for course in courses:
        code_raw = course.get("Course Code", "")
        code = str(code_raw).strip()
        if code.upper() == "HS205":
            continue
        _, T, _, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        tutorials_needed = T
        attempts = 0
        while tutorials_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(tutorial_slots)
            placed = False
            for d in days:
                if course_in_day(code + "_TUT", d):
                    continue
                for slot in tutorial_slots:
                    if timetable[d][slot] == "":
                        timetable[d][slot] = code + "_TUT"
                        tutorials_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break

    # Ensure every course in the color_map has a non-empty color.
    for c in courses:
        code_raw = c.get("Course Code", "")
        c_code = str(code_raw).strip()
        if c_code not in color_map or not color_map[c_code]:
            color_map[c_code] = "#FFD700"  # fallback gold

    return timetable

# ------------------------------
# FLASK ROUTES
# ------------------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    uploaded_file = request.files['excel_file']
    if uploaded_file.filename != '':
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
        uploaded_file.save(filepath)

        # 1) Extract color codes using openpyxl
        wb = load_workbook(filepath, data_only=True)
        sheet = wb.active

        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[cell.value] = cell.column_letter
        course_code_col = headers.get("Course Code", "B")

        color_map = {}
        for row_idx in range(2, sheet.max_row + 1):
            code_cell = sheet[f"{course_code_col}{row_idx}"]
            if code_cell.value is None:
                continue
            code_value = str(code_cell.value).strip()
            if not code_value:
                continue
            fill_color = code_cell.fill.fgColor
            if fill_color and hasattr(fill_color, 'rgb') and fill_color.rgb:
                fill_rgb = str(fill_color.rgb)
                if len(fill_rgb) == 8:
                    color_map[code_value] = f"#{fill_rgb[2:]}"
                else:
                    color_map[code_value] = f"#{fill_rgb}"
            else:
                color_map[code_value] = ""

        # 2) Read courses from Excel using pandas
        df = pd.read_excel(filepath)
        courses = df.to_dict('records')

        # 3) Generate timetable
        timetable = schedule_courses(courses, color_map)

        # 4) Build context
        global timetable_context
        timetable_context = {
            "institute_name": "Indian Institute of Information Technology Dharwad",
            "academic_year": request.form.get('academic_year', 'Jan - April 2025'),
            "semester": request.form.get('semester', 'IV'),
            "section": request.form.get('section', 'Section A'),
            "branch": "CSE",
            "group_mail": "2023csea@iiitdwd.ac.in",
            "color_map": color_map,
            "timetable": timetable,
            "courses": courses
        }
        return redirect(url_for('show_timetable'))

    return "No file selected", 400

@app.route('/timetable')
def show_timetable():
    return render_template('timetable.html', **timetable_context)

if __name__ == '__main__':
    app.run(debug=True)
