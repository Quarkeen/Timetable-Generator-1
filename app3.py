from flask import Flask, render_template, request, redirect, url_for
import os
import random
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

timetable_context = {}

def parse_credits(credit_str):
    """
    Parse a credit string of the form "L-T-P-S-C" and return (L, T, P, S, C).
    If parsing fails, return zeros.
    """
    try:
        L_str, T_str, P_str, S_str, C_str = credit_str.split('-')
        return int(L_str), int(T_str), int(P_str), int(S_str), int(C_str)
    except:
        return 0, 0, 0, 0, 0

def parse_time_24h(time_str):
    """
    Parse a single 24-hour time string like '09:00' or '17:30' -> float (9.0, 17.5, etc.)
    If invalid, return 0.0
    """
    time_str = time_str.strip()  # e.g. "09:00"
    if ':' not in time_str:
        return 0.0
    hhmm = time_str.split(':')
    if len(hhmm) != 2:
        return 0.0
    hh = int(hhmm[0])
    mm = int(hhmm[1])
    return hh + mm/60.0

def parse_time_range_24h(slot_str):
    """
    Parse a slot like '09:00 - 10:30' (24h format).
    We'll extract the start time substring and parse it with parse_time_24h.
    If we can't parse, return 0.0
    """
    slot_str = slot_str.strip()  # e.g. "09:00 - 10:30"
    parts = slot_str.split('-')
    if len(parts) != 2:
        return 0.0
    start_part = parts[0].strip()  # "09:00"
    return parse_time_24h(start_part)

def schedule_courses(courses, color_map,
                     lecture_slots, tutorial_slots, lab_slots, minor_slots,
                     morning_break, lunch_break):
    """
    1) Combine all user-provided slots (minor, lecture, morning_break, tutorial, lunch_break, lab).
    2) Sort them by start time (24-hour) so they appear in chronological order.
    3) Keep minor slots empty (not used for scheduling).
    4) Place HS205 only if there's a "17:00 - 18:30" slot, randomly in one of the 5 days.
    5) Randomly schedule L, T, P in the user-defined lecture, tutorial, lab slots.
    """
    days = ["MON", "TUE", "WED", "THU", "FRI"]
    # We'll treat "17:00 - 18:30" as the special HS205 slot if user includes it.
    hs205_slot = "17:00 - 18:30"

    # Combine all user slots in the order: minor -> lecture -> morning break -> tutorial -> lunch break -> lab
    combined_slots = []
    combined_slots.extend(minor_slots)     # minor
    combined_slots.extend(lecture_slots)   # lecture
    combined_slots.append(morning_break)   # morning break
    combined_slots.extend(tutorial_slots)  # tutorial
    combined_slots.append(lunch_break)     # lunch break
    combined_slots.extend(lab_slots)       # labs

    # If we want to ensure the HS205 slot is recognized if not already present, we can forcibly add it:
    if hs205_slot not in combined_slots:
        combined_slots.append(hs205_slot)

    # Sort by start time in 24h format
    combined_slots = sorted(combined_slots, key=lambda s: parse_time_range_24h(s))

    # Initialize timetable
    timetable = {}
    for d in days:
        timetable[d] = {}
        for slot in combined_slots:
            if slot == morning_break:
                timetable[d][slot] = "Morning Break"
            elif slot == lunch_break:
                timetable[d][slot] = "Lunch Break"
            else:
                timetable[d][slot] = ""

    def course_in_day(course_label, day):
        for val in timetable[day].values():
            if val == course_label:
                return True
        return False

    MAX_ATTEMPTS = 50

    # 1) Place HS205 in "17:00 - 18:30" if it exists
    for c in courses:
        code = str(c.get("Course Code", "")).strip().upper()
        if code == "HS205":
            if hs205_slot in combined_slots:
                # randomly place among MON-FRI
                attempts = 0
                placed = False
                while not placed and attempts < MAX_ATTEMPTS:
                    attempts += 1
                    random.shuffle(days)
                    for d in days:
                        if timetable[d][hs205_slot] == "":
                            timetable[d][hs205_slot] = "HS205"
                            placed = True
                            break
            break  # done with HS205

    # 2) Schedule labs
    for c in courses:
        code = str(c.get("Course Code", "")).strip()
        if code.upper() == "HS205":
            continue
        L, T, P, S, C = parse_credits(c.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        if P > 0:
            labs_needed = P // 2 if P >= 2 else P
            attempts = 0
            while labs_needed > 0 and attempts < MAX_ATTEMPTS:
                attempts += 1
                random.shuffle(days)
                random.shuffle(lab_slots)
                placed = False
                for d in days:
                    if course_in_day(code, d) or course_in_day(code + "_LAB", d):
                        continue
                    for lab in lab_slots:
                        if lab in timetable[d] and timetable[d][lab] == "":
                            timetable[d][lab] = code + "_LAB(2hrs)"
                            labs_needed -= 1
                            placed = True
                            break
                    if placed:
                        break

    # 3) Schedule lectures (L) in lecture_slots only
    for c in courses:
        code = str(c.get("Course Code", "")).strip()
        if code.upper() == "HS205":
            continue
        L, _, _, _, _ = parse_credits(c.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        lectures_needed = L-1
        attempts = 0
        while lectures_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(lecture_slots)
            placed = False
            for d in days:
                if course_in_day(code, d) or course_in_day(code + "_LAB", d):
                    continue
                for ls in lecture_slots:
                    if ls in timetable[d] and timetable[d][ls] == "":
                        timetable[d][ls] = code
                        lectures_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break

    # 4) Schedule tutorials (T) in tutorial_slots only
    for c in courses:
        code = str(c.get("Course Code", "")).strip()
        if code.upper() == "HS205":
            continue
        _, T, _, _, _ = parse_credits(c.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        tutorials_needed = T
        attempts = 0
        while tutorials_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(tutorial_slots)
            placed = False
            for d in days:
                if course_in_day(code + "_TUT", d):
                    continue
                for ts in tutorial_slots:
                    if ts in timetable[d] and timetable[d][ts] == "":
                        timetable[d][ts] = code + "_TUT"
                        tutorials_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break

    # Minor slots remain empty (not used for scheduling)
    # Make sure each course has a color
    for c in courses:
        c_code = str(c.get("Course Code", "")).strip()
        if c_code not in color_map or not color_map[c_code]:
            color_map[c_code] = "#FFD700"

    return timetable

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'excel_file' not in request.files:
        return "No file part", 400
    uploaded_file = request.files['excel_file']
    if uploaded_file.filename == '':
        return "No file selected", 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
    uploaded_file.save(filepath)

    # 1) Read user inputs for the various slots
    num_lecture_slots = int(request.form.get('num_lecture_slots', 2))
    lecture_slots = []
    for i in range(1, num_lecture_slots+1):
        slot_key = f"lecture_slot_{i}"
        if slot_key in request.form:
            lecture_slots.append(request.form[slot_key].strip())

    num_tutorial_slots = int(request.form.get('num_tutorial_slots', 1))
    tutorial_slots = []
    for i in range(1, num_tutorial_slots+1):
        slot_key = f"tutorial_slot_{i}"
        if slot_key in request.form:
            tutorial_slots.append(request.form[slot_key].strip())

    num_lab_slots = int(request.form.get('num_lab_slots', 1))
    lab_slots = []
    for i in range(1, num_lab_slots+1):
        slot_key = f"lab_slot_{i}"
        if slot_key in request.form:
            lab_slots.append(request.form[slot_key].strip())

    # 2) Minor slots
    num_minor_slots = int(request.form.get('num_minor_slots', 0))
    minor_slots = []
    for i in range(1, num_minor_slots+1):
        slot_key = f"minor_slot_{i}"
        if slot_key in request.form:
            minor_slots.append(request.form[slot_key].strip())

    morning_break = request.form.get('morning_break', '10:30 - 11:00').strip()
    lunch_break = request.form.get('lunch_break', '13:30 - 14:30').strip()

    # 3) Extract color codes from Excel
    wb = load_workbook(filepath, data_only=True)
    sheet = wb.active

    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[cell.value] = cell.column_letter
    course_code_col = headers.get("Course Code", "B")

    color_map = {}
    for row_idx in range(2, sheet.max_row + 1):
        code_cell = sheet[f"{course_code_col}{row_idx}"]
        if code_cell.value is None:
            continue
        code_value = str(code_cell.value).strip()
        if not code_value:
            continue
        fill_color = code_cell.fill.fgColor
        if fill_color and hasattr(fill_color, 'rgb') and fill_color.rgb:
            fill_rgb = str(fill_color.rgb)
            if len(fill_rgb) == 8:
                color_map[code_value] = f"#{fill_rgb[2:]}"
            else:
                color_map[code_value] = f"#{fill_rgb}"
        else:
            color_map[code_value] = ""

    # 4) Read courses with pandas
    df = pd.read_excel(filepath)
    courses = df.to_dict('records')

    # 5) Generate timetable
    timetable = schedule_courses(
        courses, color_map,
        lecture_slots, tutorial_slots, lab_slots, minor_slots,
        morning_break, lunch_break
    )

    # 6) Build context
    global timetable_context
    timetable_context = {
        "institute_name": "Indian Institute of Information Technology Dharwad",
        "academic_year": request.form.get('academic_year', 'Jan - April 2025'),
        "semester": request.form.get('semester', 'IV'),
        "classroom": request.form.get('classroom', 'C104'),
        "branch": request.form.get('branch', 'CSE'),
        "group_mail": request.form.get('group_mail', '2023csea@iiitdwd.ac.in'),
        "color_map": color_map,
        "timetable": timetable,
        "courses": courses
    }
    return redirect(url_for('show_timetable'))

@app.route('/timetable')
def show_timetable():
    return render_template('timetable.html', **timetable_context)

if __name__ == '__main__':
    app.run(debug=True)
