from flask import Flask, render_template, request, redirect, url_for
import os
import random
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

timetable_context = {}

# ------------------------------
# UTILITY: Parse Credits
# ------------------------------
def parse_credits(credit_str):
    """
    Parse a credit string of the form "L-T-P-S-C" and return (L, T, P, S, C).
    If parsing fails, return zeros.
    """
    try:
        L_str, T_str, P_str, S_str, C_str = credit_str.split('-')
        return int(L_str), int(T_str), int(P_str), int(S_str), int(C_str)
    except:
        return 0, 0, 0, 0, 0

# ------------------------------
# SCHEDULING FUNCTION
# ------------------------------
def schedule_courses(courses, color_map, lecture_slots, tutorial_slots, lab_slots,
                     morning_break, lunch_break):
    """
    Randomly schedule each course's L (lecture), T (tutorial), and P (lab) hours
    across 5 days (MON-FRI) using user-defined slot timings and breaks.
    
    The day is constructed by combining the provided slots and inserting
    the morning and lunch breaks as non-schedulable.
    
    The final order of slots in a day is:
      1. All lecture slots (as provided)
      2. Morning Break (user-defined)
      3. All tutorial slots (as provided)
      4. Lunch Break (user-defined)
      5. All lab slots (as provided)
      6. A reserved slot for HS205 ("5:00 - 6:30 PM")
    """
    days = ["MON", "TUE", "WED", "THU", "FRI"]

    # Build the day's slot order:
    all_slots = []
    # Add lecture slots
    for ls in lecture_slots:
        all_slots.append(ls)
    # Insert morning break
    all_slots.append(morning_break)
    # Add tutorial slots
    for ts in tutorial_slots:
        all_slots.append(ts)
    # Insert lunch break
    all_slots.append(lunch_break)
    # Add lab slots
    for lab in lab_slots:
        all_slots.append(lab)
    # Add reserved HS205 slot (hardcoded)
    hs205_slot = "5:00 - 6:30 PM"
    all_slots.append(hs205_slot)

    # Initialize timetable: dictionary of day -> {slot: value}
    timetable = {}
    for d in days:
        timetable[d] = {}
        for slot in all_slots:
            if slot == morning_break:
                timetable[d][slot] = "Morning Break"
            elif slot == lunch_break:
                timetable[d][slot] = "Lunch Break"
            else:
                timetable[d][slot] = ""

    # Helper: check if a course (or variant) is already scheduled in a day.
    def course_in_day(course_label, day):
        for val in timetable[day].values():
            if val == course_label:
                return True
        return False

    MAX_ATTEMPTS = 50

    # 1) Reserve HS205 in the reserved slot on FRIDAY
    for course in courses:
        if str(course.get("Course Code", "")).strip().upper() == "HS205":
            timetable["FRI"][hs205_slot] = "HS205"
            break

    # 2) Schedule lab sessions for courses with P > 0.
    for course in courses:
        code = str(course.get("Course Code", "")).strip()
        # Skip HS205, which is already placed
        if code.upper() == "HS205":
            continue
        _, _, P, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        if P > 0:
            labs_needed = P // 2 if P >= 2 else P
            attempts = 0
            while labs_needed > 0 and attempts < MAX_ATTEMPTS:
                attempts += 1
                random.shuffle(days)
                random.shuffle(lab_slots)
                placed = False
                for d in days:
                    # Avoid duplicating lab on the same day
                    if course_in_day(code + "_LAB", d) or course_in_day(code, d):
                        continue
                    for lab in lab_slots:
                        if timetable[d][lab] == "":
                            timetable[d][lab] = code + "_LAB(2hrs)"
                            labs_needed -= 1
                            placed = True
                            break
                    if placed:
                        break
            # End lab allocation for this course

    # 3) Schedule lecture sessions (L)
    for course in courses:
        code = str(course.get("Course Code", "")).strip()
        if code.upper() == "HS205":
            continue
        L, _, _, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        lectures_needed = L-1
        attempts = 0
        while lectures_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(lecture_slots)
            placed = False
            for d in days:
                # Ensure same course doesn't appear twice in the same day (lecture variant)
                if course_in_day(code, d):
                    continue
                for ls in lecture_slots:
                    if timetable[d][ls] == "":
                        timetable[d][ls] = code
                        lectures_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break

    # 4) Schedule tutorial sessions (T)
    for course in courses:
        code = str(course.get("Course Code", "")).strip()
        if code.upper() == "HS205":
            continue
        _, T, _, _, _ = parse_credits(course.get("Credits (L-T-P-S-C)", "0-0-0-0-0"))
        tutorials_needed = T
        attempts = 0
        while tutorials_needed > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            random.shuffle(days)
            random.shuffle(tutorial_slots)
            placed = False
            for d in days:
                if course_in_day(code + "_TUT", d):
                    continue
                for ts in tutorial_slots:
                    if timetable[d][ts] == "":
                        timetable[d][ts] = code + "_TUT"
                        tutorials_needed -= 1
                        placed = True
                        break
                if placed:
                    break
            if not placed:
                break

    # Ensure every course has a color; fallback to gold (#FFD700)
    for c in courses:
        c_code = str(c.get("Course Code", "")).strip()
        if c_code not in color_map or not color_map[c_code]:
            color_map[c_code] = "#FFD700"

    return timetable

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'excel_file' not in request.files:
        return "No file part", 400
    uploaded_file = request.files['excel_file']
    if uploaded_file.filename == '':
        return "No file selected", 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
    uploaded_file.save(filepath)

    # 1) Read slot constraints from user input
    num_lecture_slots = int(request.form.get('num_lecture_slots', 2))
    lecture_slots = []
    for i in range(1, num_lecture_slots+1):
        slot_key = f"lecture_slot_{i}"
        if slot_key in request.form:
            lecture_slots.append(request.form[slot_key].strip())

    num_tutorial_slots = int(request.form.get('num_tutorial_slots', 1))
    tutorial_slots = []
    for i in range(1, num_tutorial_slots+1):
        slot_key = f"tutorial_slot_{i}"
        if slot_key in request.form:
            tutorial_slots.append(request.form[slot_key].strip())

    num_lab_slots = int(request.form.get('num_lab_slots', 1))
    lab_slots = []
    for i in range(1, num_lab_slots+1):
        slot_key = f"lab_slot_{i}"
        if slot_key in request.form:
            lab_slots.append(request.form[slot_key].strip())

    morning_break = request.form.get('morning_break', '10:30 - 11:00 AM')
    lunch_break = request.form.get('lunch_break', '1:30 - 2:30 PM')

    # 2) Extract color codes from Excel using openpyxl
    wb = load_workbook(filepath, data_only=True)
    sheet = wb.active

    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[cell.value] = cell.column_letter
    course_code_col = headers.get("Course Code", "B")

    color_map = {}
    for row_idx in range(2, sheet.max_row + 1):
        code_cell = sheet[f"{course_code_col}{row_idx}"]
        if code_cell.value is None:
            continue
        code_value = str(code_cell.value).strip()
        if not code_value:
            continue
        fill_color = code_cell.fill.fgColor
        if fill_color and hasattr(fill_color, 'rgb') and fill_color.rgb:
            fill_rgb = str(fill_color.rgb)
            if len(fill_rgb) == 8:
                color_map[code_value] = f"#{fill_rgb[2:]}"
            else:
                color_map[code_value] = f"#{fill_rgb}"
        else:
            color_map[code_value] = ""

    # 3) Read courses from Excel using pandas
    df = pd.read_excel(filepath)
    courses = df.to_dict('records')

    # 4) Generate timetable using user-defined slot constraints
    timetable = schedule_courses(courses, color_map, lecture_slots, tutorial_slots, lab_slots,
                                 morning_break, lunch_break)

    # 5) Build context
    global timetable_context
    timetable_context = {
        "institute_name": "Indian Institute of Information Technology Dharwad",
        "academic_year": request.form.get('academic_year', 'Jan - April 2025'),
        "semester": request.form.get('semester', 'IV'),
        "section": request.form.get('section', 'Section A'),
        "branch": request.form.get('branch', 'CSE'),
        "group_mail": request.form.get('group_mail', '2023csea@iiitdwd.ac.in'),
        "color_map": color_map,
        "timetable": timetable,
        "courses": courses
    }
    return redirect(url_for('show_timetable'))

@app.route('/timetable')
def show_timetable():
    return render_template('timetable.html', **timetable_context)

if __name__ == '__main__':
    app.run(debug=True)
